import streamlit as st
import base64
import json
import re
import io
import copy
import math
from datetime import datetime
from PIL import Image
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
VALID_ROOMS = (
    [f"卡{i}" for i in range(1, 7)]
    + ["111", "222", "333", "555", "666", "777", "888", "999"]
    + [f"厅{i}" for i in range(1, 21)]
)

FEE_RATES = {
    "现金": 0, "会员卡": 0, "挂帐": 0,
    "微信": 0.0038, "支付宝": 0.0038, "收钱吧": 0.0038,
    "抖音": 0.06,
    "饿了么": 0, "美团": 0,
}

REPORT_PROMPT = """你是一个手写餐厅日报 OCR 助手。请仔细识别这张手写日报照片中的所有信息。

**照片结构说明**：
- 右上角有日期（年/月/日）
- 表格从左到右的列依次是：序号、用餐时间(中/晚)、包间号、营业额、(空列)、收入、付款方式、实际收款
- 表格下方可能有备注信息

**包间号只可能是以下值之一**：
卡1, 卡2, 卡3, 卡4, 卡5, 卡6, 111, 222, 333, 555, 666, 777, 888, 999, 厅1~厅20

**付款方式只可能是以下值**：
支付宝, 抖音, 微信, 现金, 饿了么, 美团, 收钱吧, 会员卡

**【重点】付款方式列的识别规则**：
- 如果只有一种付款方式且没有金额数字，说明整单都是这一种方式，金额 = 收入列的金额
- 如果付款方式列写了多种来源（如"抖音500 微信300"或"微信/现金200"），则每种方式后面跟的数字就是该方式的实际收到金额
- **关键校验**：所有付款方式的金额相加 必须等于 收入列的金额
- **关键校验**：收入列的金额 不能超过 营业额
- 请仔细辨认付款方式后面跟的数字，这些数字容易和收入列的数字混淆

请以严格的 JSON 格式输出，结构如下：
{
  "date": "YYYY-MM-DD",
  "rows": [
    {
      "seq": 1,
      "period": "中或晚",
      "period_uncertain": false,
      "room": "包间号",
      "room_uncertain": false,
      "revenue": 数字(营业额,整单总营业额),
      "revenue_uncertain": false,
      "income": 数字(收入列的金额),
      "payments": [
        {"method": "微信", "amount": 500, "uncertain": false},
        {"method": "现金", "amount": 200, "uncertain": false}
      ],
      "drinks": 0,
      "row_note": ""
    }
  ],
  "notes": "底部备注内容，没有则为空字符串"
}

**重要规则**：
1. 只提取有实际数据的行，空行跳过
2. 对于你不确定的字段，将对应的 xxx_uncertain 设为 true
3. 如果某个字段完全看不清，填 null 并标记 uncertain 为 true
4. 日期从右上角识别，格式为 YYYY-MM-DD
5. 如果一笔订单只有一种付款方式，payments 数组只有一个元素，amount = income
6. 如果有多种付款方式，分别列出每种方式和金额，所有 amount 之和必须 = income
7. income 不能超过 revenue
8. drinks 是酒水列的金额，没有则为 0
9. row_note 是该行末尾的备注（如"未付"、"陈姐未付"等），没有则为空字符串
10. 只输出 JSON，不要输出任何其他文字"""

# ---------------------------------------------------------------------------
# 样式常量
# ---------------------------------------------------------------------------
_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
_CENTER = Alignment(horizontal="center", vertical="center")
_HFILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HFONT = Font(bold=True, size=10, color="FFFFFF")
_BFONT = Font(bold=True, size=10)
_MFMT = '#,##0.00'

# 匹配模板：A-R 共 18 列
DAY_HEADERS = [
    "序号", "中餐/晚餐", "包间号", "营业额", "折扣", "收入",
    "手续费", "饿了么", "实收", "挂账", "挂账收回",
    "会员卡赠送", "会员卡消费", "陈其兵", "付款方式", "酒水", "备注", "充值",
]
DAY_WIDTHS = [6, 10, 8, 10, 8, 10, 10, 8, 10, 8, 10, 10, 10, 10, 12, 8, 14, 8]

# 汇总表列头匹配模板
SUMMARY_HEADERS = [
    "序号", "日期", "营业额", "折扣", "收入", "手续费", "饿了么",
    "实收", "挂账", "挂账收回", "会员卡赠送", "会员卡消费", "陈其兵", "酒水",
]
SUMMARY_WIDTHS = [6, 12, 12, 10, 12, 10, 8, 12, 10, 10, 12, 12, 12, 10]

# 汇总表列 → 日sheet row28 列字母（D=营业额, E=折扣, ..., P=酒水）
_SUMMARY_DAY_COL_MAP = {
    3: "D", 4: "E", 5: "F", 6: "G", 7: "H", 8: "I",
    9: "J", 10: "K", 11: "L", 12: "M", 13: "N", 14: "P",
}


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _img_b64(image):
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def _call_vl(client, b64, prompt):
    c = client.chat.completions.create(
        model="qwen-vl-max",
        messages=[{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
            {"type": "text", "text": prompt},
        ]}],
    )
    return c.choices[0].message.content


def _parse_json(raw):
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```\w*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            return json.loads(m.group())
        raise ValueError(f"无法解析:\n{raw[:500]}")


def _ceil2(x):
    """向上进位到两位小数（0.237 → 0.24）。"""
    return math.ceil(x * 100) / 100


def _get_fee_rate(method: str) -> float:
    method = method.strip()
    return FEE_RATES.get(method, 0.0038)


def _process_notes(day_data):
    """后处理：根据行备注调整支付方式，汇总备注到底部。"""
    collected_notes = []
    for row in day_data.get("rows", []):
        note = (row.get("row_note") or "").strip()
        drinks = row.get("drinks") or 0
        remark_parts = []

        if note:
            if "陈姐未付" in note:
                row["payments"] = [{"method": "现金", "amount": row.get("income") or 0}]
                remark_parts.append("陈姐未付")
            elif "未付" in note:
                row["payments"] = [{"method": "挂帐", "amount": row.get("income") or 0}]
                remark_parts.append("未付")
            else:
                remark_parts.append(note)

        if drinks:
            remark_parts.append(f"酒水{drinks}")

        row["row_note"] = "；".join(remark_parts)

        if remark_parts:
            seq = row.get("seq", "?")
            collected_notes.append(f"第{seq}单：{'，'.join(remark_parts)}")

    if collected_notes:
        existing = (day_data.get("notes") or "").strip()
        combined = "；".join(collected_notes)
        day_data["notes"] = f"{existing}；{combined}".strip("；") if existing else combined


def _validate_row(row):
    """校验一行数据，返回警告列表。"""
    warnings = []
    revenue = row.get("revenue") or 0
    income = row.get("income") or 0
    payments = row.get("payments", [])
    pay_sum = sum(p.get("amount") or 0 for p in payments)

    if payments and abs(pay_sum - income) > 0.5:
        warnings.append(f"付款合计({pay_sum}) ≠ 收入({income})")
    if income > revenue > 0:
        warnings.append(f"收入({income}) > 营业额({revenue})")
    return warnings


def _validate_day(day_data):
    """校验一天的数据，将警告写入每行的 _warnings 字段，返回是否有问题。"""
    has_issue = False
    for row in day_data.get("rows", []):
        row["_warnings"] = _validate_row(row)
        if row["_warnings"]:
            has_issue = True
    return has_issue


def _has_uncertain(day_data):
    for row in day_data.get("rows", []):
        if row.get("_warnings"):
            return True
        for key in row:
            if key.endswith("_uncertain") and row[key]:
                return True
        for p in row.get("payments", []):
            if p.get("uncertain"):
                return True
    return False


def _sheet_has_data(ws):
    """判断一个日 sheet 是否有实际流水数据（D列 row 3-27 有数字）。"""
    for r in range(3, 28):
        v = ws.cell(row=r, column=4).value
        if v is not None and isinstance(v, (int, float)) and v > 0:
            return True
    return False


def _flatten_rows_for_excel(rows):
    """将带 payments 数组的行展开为每个支付方式一行。同一单序号相同。"""
    flat = []
    for row in rows:
        payments = row.get("payments", [])
        if not payments:
            payments = [{"method": "", "amount": 0}]
        seq = row.get("seq", 0)
        period = row.get("period", "")
        room = row.get("room", "")
        revenue = row.get("revenue") or 0
        row_income = row.get("income") or 0
        drinks = row.get("drinks") or 0
        row_note = row.get("row_note") or ""
        is_first = True
        for pay in payments:
            method = pay.get("method", "")
            amount = pay.get("amount") or 0
            is_member = method == "会员卡"
            is_guazhang = method == "挂帐"
            fee_rate = _get_fee_rate(method)
            income = 0 if (is_member or is_guazhang) else amount
            fee = _ceil2(income * fee_rate)
            actual = _ceil2(income - fee)

            flat.append({
                "seq": seq,
                "period": period,
                "room": room,
                "revenue": revenue if is_first else 0,
                "row_income": row_income if is_first else 0,
                "income": income,
                "fee": fee,
                "actual": actual,
                "payment": method,
                "amount": amount,
                "is_first": is_first,
                "drinks": drinks if is_first else 0,
                "row_note": row_note if is_first else "",
            })
            is_first = False
    return flat


# ---------------------------------------------------------------------------
# 合并单元格安全处理
# ---------------------------------------------------------------------------
# 设计目标：真正兼容用户各种合并单元格模板，而非简单跳过。
#
# 诊断级别（严重性由低到高）：
#   INFO    清空操作落在合并块上，行为等价（清锚点 == 清整块）
#   NOTICE  单次写入被重定向到合并块锚点，数据未丢失但格式需用户确认
#   CONFLICT 多条数据竞争同一个合并锚点，仅保留首条（选项 1 策略）
#   REJECT  合并块跨越数据区边界（如 B2:B5 跨表头和数据区），为保护模板完全不动

_DIAG_INFO = "INFO"
_DIAG_NOTICE = "NOTICE"
_DIAG_CONFLICT = "CONFLICT"
_DIAG_REJECT = "REJECT"

_DATA_ROW_MIN = 3   # 数据区第一行
_DATA_ROW_MAX = 27  # 数据区最后一行
_DATA_COL_MIN = 2   # 数据区第一列 (B)
_DATA_COL_MAX = 18  # 数据区最后一列 (R)


def _build_merge_map(ws):
    """构建 (row,col) -> 合并信息 的映射。
    返回:
        merge_map: {(r,c): {"anchor": (ar,ac), "range": "B3:D3",
                            "min_row","max_row","min_col","max_col",
                            "crosses_boundary": bool}}
    """
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        ar, ac = mr.min_row, mr.min_col
        range_str = str(mr)
        crosses = (
            mr.min_row < _DATA_ROW_MIN or mr.max_row > _DATA_ROW_MAX
            or mr.min_col < _DATA_COL_MIN or mr.max_col > _DATA_COL_MAX
        )
        info = {
            "anchor": (ar, ac),
            "range": range_str,
            "min_row": mr.min_row, "max_row": mr.max_row,
            "min_col": mr.min_col, "max_col": mr.max_col,
            "crosses_boundary": crosses,
        }
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = info
    return merge_map


def _cell_ref(r, c):
    return f"{get_column_letter(c)}{r}"


def _safe_set(ws, merge_map, r, c, value, *, diagnostics, sheet_name,
              field_desc="", written_anchors=None):
    """安全写入：自动处理合并单元格。

    written_anchors: set，记录本轮已写过的锚点；用于检测 CONFLICT。
                     由调用方（每个数据行写入前）维护。
    field_desc: 字段说明，用于诊断信息可读性（如"序号"/"房号"）。
    返回: True 正常写入 / False 被合并规则拒绝。
    """
    info = merge_map.get((r, c))
    target_ref = _cell_ref(r, c)

    if info is None:
        # 普通 Cell，直接写
        ws.cell(row=r, column=c, value=value)
        return True

    # 跨边界合并：拒绝写入，保护模板
    if info["crosses_boundary"]:
        diagnostics.append({
            "level": _DIAG_REJECT,
            "sheet": sheet_name,
            "range": info["range"],
            "target": target_ref,
            "field": field_desc,
            "value": value,
            "msg": f"合并区 {info['range']} 跨越表头/数据区边界，为保护模板未写入。",
        })
        return False

    anchor = info["anchor"]
    anchor_ref = _cell_ref(*anchor)

    # 冲突检测：同一轮已经写过这个锚点
    if written_anchors is not None and anchor in written_anchors:
        diagnostics.append({
            "level": _DIAG_CONFLICT,
            "sheet": sheet_name,
            "range": info["range"],
            "target": target_ref,
            "anchor": anchor_ref,
            "field": field_desc,
            "value": value,
            "msg": (f"合并区 {info['range']} 的锚点 {anchor_ref} 已被本行/前一条数据占用，"
                    f"当前值「{value}」未能写入（只保留首条数据）。"),
        })
        return False

    # 重定向到锚点写入
    ws.cell(row=anchor[0], column=anchor[1], value=value)
    if written_anchors is not None:
        written_anchors.add(anchor)

    # 位置重定向了才告警；锚点就是目标本身的话属于正常情况
    if (r, c) != anchor:
        diagnostics.append({
            "level": _DIAG_NOTICE,
            "sheet": sheet_name,
            "range": info["range"],
            "target": target_ref,
            "anchor": anchor_ref,
            "field": field_desc,
            "value": value,
            "msg": (f"目标单元格 {target_ref} 属于合并区 {info['range']}，"
                    f"已改写入锚点 {anchor_ref}。"),
        })
    return True


def _safe_clear_range(ws, merge_map, r1, c1, r2, c2, *, diagnostics, sheet_name):
    """安全清空指定矩形区域。
    - 普通 Cell 置空
    - 合并块：若完全落在数据区内，只清锚点（等价清整块），记 INFO
    - 合并块：若跨边界，完全不动，记 REJECT
    - 合并块非锚点位置不重复记录
    """
    seen_anchors = set()
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            info = merge_map.get((r, c))
            if info is None:
                ws.cell(row=r, column=c, value=None)
                continue

            anchor = info["anchor"]
            if anchor in seen_anchors:
                continue
            seen_anchors.add(anchor)

            if info["crosses_boundary"]:
                diagnostics.append({
                    "level": _DIAG_REJECT,
                    "sheet": sheet_name,
                    "range": info["range"],
                    "target": _cell_ref(r, c),
                    "msg": f"清空时遇到跨边界合并区 {info['range']}，已保留原值未清空。",
                })
                continue

            # 合并块在数据区内：清锚点即清整块
            ar, ac = anchor
            ws.cell(row=ar, column=ac, value=None)
            diagnostics.append({
                "level": _DIAG_INFO,
                "sheet": sheet_name,
                "range": info["range"],
                "anchor": _cell_ref(ar, ac),
                "msg": f"清空了合并区 {info['range']}（通过清锚点 {_cell_ref(ar, ac)} 实现）。",
            })


def _fill_day_data(ws, day_data, *, diagnostics=None, sheet_name=""):
    """往已有模板 sheet 中填入数据（只写数据行 3-27，保留表头和合计公式）。

    兼容合并单元格：
    - 清空数据区时跳过跨边界合并块，对纯数据区内合并块只清锚点
    - 写入数据时自动重定向到合并块锚点，多行竞争同锚点时仅保留首条并告警
    """
    if diagnostics is None:
        diagnostics = []  # 独立调用时不抛异常，只是不收集
    merge_map = _build_merge_map(ws)

    # 清空数据区 B3:R27（安全方式）
    _safe_clear_range(ws, merge_map,
                      _DATA_ROW_MIN, _DATA_COL_MIN,
                      _DATA_ROW_MAX, _DATA_COL_MAX,
                      diagnostics=diagnostics, sheet_name=sheet_name)

    flat = _flatten_rows_for_excel(day_data.get("rows", []))
    current_row = _DATA_ROW_MIN

    for entry in flat:
        if current_row > _DATA_ROW_MAX:
            break
        r = current_row
        current_row += 1

        # 每个数据行用独立的 written_anchors，避免跨行误报冲突
        written = set()

        def put(col, value, field):
            _safe_set(ws, merge_map, r, col, value,
                      diagnostics=diagnostics, sheet_name=sheet_name,
                      field_desc=field, written_anchors=written)

        put(1, entry["seq"], "序号")
        put(2, entry["period"], "时段")
        put(3, entry["room"], "房号")

        if entry["is_first"] and entry["revenue"]:
            put(4, entry["revenue"], "营业额")

        if entry["is_first"] and entry["revenue"] and entry["row_income"]:
            discount = _ceil2(entry["revenue"] - entry["row_income"])
            if discount > 0:
                put(5, discount, "折扣")

        put(6, entry["income"], "收入")
        put(7, entry["fee"], "手续费")
        put(9, f"=F{r}-G{r}", "实收公式")
        put(15, entry["payment"], "付款方式")

        if entry.get("drinks"):
            put(16, entry["drinks"], "酒水")
        if entry.get("row_note"):
            put(17, entry["row_note"], "备注")

    notes = day_data.get("notes", "")
    if notes:
        # 备注写在 J29（非数据区），这里不需要合并单元格处理；但为稳妥也走 safe_set
        _safe_set(ws, merge_map, 29, 10, f"备注：{notes}",
                  diagnostics=diagnostics, sheet_name=sheet_name,
                  field_desc="日备注")


def _write_day_sheet(ws, day_data):
    """从零创建一个日结 sheet（用于没有模板的情况）。"""
    date_str = day_data.get("date", "")
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        ws["K1"] = f"日期：  {dt.year}  年  {dt.month}  月  {dt.day}  日"
    except Exception:
        ws["K1"] = f"日期：{date_str}"
    ws["K1"].font = _BFONT

    for ci, (h, w) in enumerate(zip(DAY_HEADERS, DAY_WIDTHS), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = _HFONT, _HFILL, _CENTER, _BORDER
        ws.column_dimensions[c.column_letter].width = w

    for ri in range(3, 28):
        ws.cell(row=ri, column=1, value=ri - 2).border = _BORDER
        ws.cell(row=ri, column=1).alignment = _CENTER
        for ci in range(2, 19):
            ws.cell(row=ri, column=ci).border = _BORDER
            ws.cell(row=ri, column=ci).alignment = _CENTER

    flat = _flatten_rows_for_excel(day_data.get("rows", []))
    current_row = 3

    for entry in flat:
        if current_row > 27:
            break
        r = current_row
        current_row += 1

        ws.cell(row=r, column=1, value=entry["seq"])
        ws.cell(row=r, column=2, value=entry["period"])
        ws.cell(row=r, column=3, value=entry["room"])

        if entry["is_first"] and entry["revenue"]:
            ws.cell(row=r, column=4, value=entry["revenue"]).number_format = _MFMT

        if entry["is_first"] and entry["revenue"] and entry["row_income"]:
            discount = _ceil2(entry["revenue"] - entry["row_income"])
            if discount > 0:
                ws.cell(row=r, column=5, value=discount).number_format = _MFMT

        ws.cell(row=r, column=6, value=entry["income"]).number_format = _MFMT
        ws.cell(row=r, column=7, value=entry["fee"]).number_format = _MFMT
        ws.cell(row=r, column=9, value=f"=F{r}-G{r}")
        ws.cell(row=r, column=9).number_format = _MFMT
        ws.cell(row=r, column=15, value=entry["payment"])

        if entry.get("drinks"):
            ws.cell(row=r, column=16, value=entry["drinks"]).number_format = _MFMT
        if entry.get("row_note"):
            ws.cell(row=r, column=17, value=entry["row_note"])

    # Row 28: 合计
    ws.cell(row=28, column=1, value="合计").font = _BFONT
    ws.cell(row=28, column=1).border = _BORDER
    ws.cell(row=28, column=1).alignment = _CENTER
    for ci in range(2, 19):
        cl = get_column_letter(ci)
        c = ws.cell(row=28, column=ci)
        c.value = f"=SUM({cl}3:{cl}27)"
        c.number_format, c.border, c.alignment, c.font = _MFMT, _BORDER, _CENTER, _BFONT

    # Rows 29-38
    ws.cell(row=29, column=1, value="营业收入：").font = _BFONT
    ws.cell(row=29, column=6, value="采购支出：").font = _BFONT
    ws.cell(row=30, column=1, value="收钱吧")
    ws.cell(row=30, column=2, value="微信/支付宝")
    ws.cell(row=30, column=6, value="食材（肉，水产，蔬菜，调味品，饮料）")
    ws.cell(row=32, column=2, value="现金")
    ws.cell(row=33, column=2, value="抖音/团券")
    ws.cell(row=33, column=6, value="其它费用")
    ws.cell(row=34, column=2, value="会员卡")
    ws.cell(row=34, column=6, value="设备、维修")
    ws.cell(row=35, column=6, value="水电燃气费")
    ws.cell(row=36, column=6, value="工资")
    ws.cell(row=37, column=6, value="营销")
    ws.cell(row=38, column=1, value="合计").font = _BFONT
    ws.cell(row=38, column=3, value="=SUM(C30:C36)")

    notes = day_data.get("notes", "")
    if notes:
        ws.cell(row=29, column=10, value=f"备注：{notes}")


def _write_summary_sheet(ws, day_sheets: list[str]):
    """生成汇总 sheet，引用各日 sheet 的合计行。"""
    for ci, (h, w) in enumerate(zip(SUMMARY_HEADERS, SUMMARY_WIDTHS), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = _HFONT, _HFILL, _CENTER, _BORDER
        ws.column_dimensions[c.column_letter].width = w

    sorted_days = sorted(day_sheets, key=lambda x: int(x) if x.isdigit() else 99)

    for idx, day_num_str in enumerate(sorted_days):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx + 1).border = _BORDER
        ws.cell(row=row, column=1).alignment = _CENTER
        ws.cell(row=row, column=2, value=int(day_num_str) if day_num_str.isdigit() else day_num_str)
        ws.cell(row=row, column=2).border = _BORDER
        ws.cell(row=row, column=2).alignment = _CENTER

        for summary_col, day_col_letter in _SUMMARY_DAY_COL_MAP.items():
            c = ws.cell(row=row, column=summary_col)
            c.value = f"='{day_num_str}'!{day_col_letter}28"
            c.number_format = _MFMT
            c.border = _BORDER
            c.alignment = _CENTER

    # 合计行 (row 39 位置，匹配模板)
    total_row = len(sorted_days) + 4
    ws.cell(row=total_row, column=1, value="合计").font = _BFONT
    for ci in range(3, 15):
        cl = get_column_letter(ci)
        c = ws.cell(row=total_row, column=ci)
        c.value = f"=SUM({cl}4:{cl}{total_row - 1})"
        c.number_format = _MFMT
        c.border = _BORDER
        c.font = _BFONT


def _write_diagnostic_sheet(wb, diagnostics):
    """把收集到的诊断信息写入一个新的 sheet "_诊断说明"。
    面向普通用户：用简明中文告诉用户哪些位置因合并单元格而被特殊处理，
    请用户人工核对。
    """
    # 去掉旧的诊断 sheet（如果存在），避免重复
    if "_诊断说明" in wb.sheetnames:
        del wb["_诊断说明"]
    ws = wb.create_sheet("_诊断说明")

    # 样式
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_font = Font(bold=True, color="C00000")
    bold = Font(bold=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 55

    ws["A1"] = "合并单元格处理诊断报告"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = ("说明：模板中存在合并单元格时，程序会根据规则安全处理。"
                "请核对下列位置的内容是否符合预期。")
    ws.merge_cells("A2:F2")

    if not diagnostics:
        ws["A4"] = "本次生成未发现合并单元格相关处理情况，所有数据已按标准方式写入。"
        return

    # 级别说明
    legend = [
        ("级别图例：", ""),
        ("ℹ️ 信息", "清空操作落在合并块上，已通过清锚点实现等价清空。"),
        ("⚠️ 注意", "写入位置被重定向到合并块锚点，数据未丢失但格式需确认。"),
        ("❌ 冲突", "多条数据竞争同一合并锚点，仅保留首条，其余未写入。"),
        ("🚫 拒绝", "合并块跨越表头/数据区边界，为保护模板完全未动。"),
    ]
    for i, (a, b) in enumerate(legend, start=4):
        ws.cell(row=i, column=1, value=a).font = bold
        ws.cell(row=i, column=2, value=b)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
    for i in range(5, 9):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    # 表头
    header_row = 10
    headers = ["级别", "所在 Sheet", "合并范围", "目标格", "锚点格", "说明"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = bold
        c.fill = yellow

    # 排序：REJECT/CONFLICT 优先展示，便于用户一眼看到需要处理的问题
    level_order = {_DIAG_REJECT: 0, _DIAG_CONFLICT: 1,
                   _DIAG_NOTICE: 2, _DIAG_INFO: 3}
    level_label = {
        _DIAG_INFO: "ℹ️ 信息",
        _DIAG_NOTICE: "⚠️ 注意",
        _DIAG_CONFLICT: "❌ 冲突",
        _DIAG_REJECT: "🚫 拒绝",
    }
    sorted_diags = sorted(
        diagnostics,
        key=lambda d: (level_order.get(d["level"], 99),
                       d.get("sheet", ""), d.get("range", ""))
    )

    row = header_row + 1
    for d in sorted_diags:
        lvl = d["level"]
        ws.cell(row=row, column=1, value=level_label.get(lvl, lvl))
        ws.cell(row=row, column=2, value=d.get("sheet", ""))
        ws.cell(row=row, column=3, value=d.get("range", ""))
        ws.cell(row=row, column=4, value=d.get("target", ""))
        ws.cell(row=row, column=5, value=d.get("anchor", ""))
        ws.cell(row=row, column=6, value=d.get("msg", ""))
        if lvl in (_DIAG_CONFLICT, _DIAG_REJECT):
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).font = red_font
        row += 1

    # 统计汇总
    row += 1
    counts = {}
    for d in diagnostics:
        counts[d["level"]] = counts.get(d["level"], 0) + 1
    ws.cell(row=row, column=1, value="统计：").font = bold
    summary_text = "  ".join(
        f"{level_label.get(k, k)}: {counts[k]}" for k in
        [_DIAG_REJECT, _DIAG_CONFLICT, _DIAG_NOTICE, _DIAG_INFO] if k in counts
    )
    ws.cell(row=row, column=2, value=summary_text)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)


def _highlight_problem_cells(wb, diagnostics):
    """对 CONFLICT / REJECT 级别的问题位置，在对应的日 sheet 中用黄底 + 红字标记，
    方便用户直接在日报表上看到需要确认的格子。
    """
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_font = Font(bold=True, color="C00000")
    for d in diagnostics:
        if d["level"] not in (_DIAG_CONFLICT, _DIAG_REJECT):
            continue
        sheet = d.get("sheet")
        if not sheet or sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # 锚点优先（合并块只能在锚点上改样式），否则用目标格
        target = d.get("anchor") or d.get("target")
        if not target:
            continue
        try:
            cell = ws[target]
            cell.fill = yellow
            if cell.font:
                cell.font = Font(
                    name=cell.font.name, size=cell.font.size,
                    bold=True, color="C00000",
                )
            else:
                cell.font = red_font
        except Exception:
            # 某些合并格不能直接设样式，静默忽略
            pass


def _build_excel(all_days, existing_wb=None):
    """生成完整 Excel，支持合并已有工作簿。

    兼容合并单元格模板：所有写入/清空走安全路径，并收集诊断信息，
    最终在工作簿末尾追加一个 "_诊断说明" sheet 给用户核对。
    """
    if existing_wb:
        wb = existing_wb
    else:
        wb = Workbook()
        wb.remove(wb.active)

    diagnostics = []

    for day_data in sorted(all_days, key=lambda d: d.get("date", "")):
        date_str = day_data.get("date", "")
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            sheet_name = str(dt.day)
        except Exception:
            sheet_name = date_str or "未知"

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if _sheet_has_data(ws):
                continue
            _fill_day_data(ws, day_data,
                           diagnostics=diagnostics, sheet_name=sheet_name)
        else:
            ws = wb.create_sheet(sheet_name)
            _write_day_sheet(ws, day_data)

    # 生成/更新汇总表（不删除已有的，只更新公式引用）
    if "汇总" in wb.sheetnames:
        del wb["汇总"]
    ws_summary = wb.create_sheet("汇总", 0)

    day_sheets = [s for s in wb.sheetnames
                  if s not in ("汇总", "_诊断说明") and s.isdigit()]
    _write_summary_sheet(ws_summary, day_sheets)

    # 对冲突/拒绝项在日报表上直接高亮
    _highlight_problem_cells(wb, diagnostics)

    # 追加诊断 sheet（放在最后）
    _write_diagnostic_sheet(wb, diagnostics)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ========== UI 入口 ==========

def run():
    st.title("📋 餐厅日报表")
    st.caption("上传手写日报照片，AI 识别后生成标准 Excel 日报表")

    st.markdown("""<style>
    [data-testid="stFileUploaderDropzoneInstructions"] div span {
        visibility: hidden; position: relative;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] div span::after {
        content: "将文件拖拽到此处"; visibility: visible;
        position: absolute; left: 0; right: 0; text-align: center;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] div small {
        visibility: hidden; position: relative;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] div small::after {
        visibility: visible; position: absolute; left: 0; right: 0;
        text-align: center;
    }
    [data-testid="stFileUploaderDropzone"] button {
        visibility: hidden; position: relative;
    }
    [data-testid="stFileUploaderDropzone"] button::after {
        content: "选择文件"; visibility: visible;
        position: absolute; left: 0; right: 0; text-align: center;
    }
    section[data-testid="stFileUploader"] small {
        visibility: hidden; position: relative;
    }
    section[data-testid="stFileUploader"] small::after {
        visibility: visible; position: absolute; left: 0;
    }
    </style>""", unsafe_allow_html=True)

    if st.button("⬅️ 返回主页"):
        st.session_state.current_page = "home"
        st.rerun()

    api_key = st.secrets.get("DASHSCOPE_API_KEY", "") if hasattr(st, "secrets") else ""
    if not api_key:
        api_key = st.text_input("请输入阿里云百炼 API Key（sk- 开头）", type="password")
    if not api_key:
        st.info("请先配置 API Key 才能使用。")
        st.stop()

    if "restaurant_results" not in st.session_state:
        st.session_state.restaurant_results = None
    if "restaurant_confirmed" not in st.session_state:
        st.session_state.restaurant_confirmed = False
    if "restaurant_existing_wb" not in st.session_state:
        st.session_state.restaurant_existing_wb = None

    # --- 上传区 ---
    st.markdown("#### 上传文件")

    existing_excel = st.file_uploader(
        "上传已有的报表 Excel（可选，用于合并）",
        type=["xlsx"],
        help="之前的报表 Excel，新识别的日期数据会填入空白天的sheet，已有流水数据的天不会被覆盖",
    )

    uploaded_photos = st.file_uploader(
        "上传手写日报照片（可多选）",
        type=["png", "jpg", "jpeg"],
        accept_multiple_files=True,
        help="每张照片对应一天的手写日报",
    )

    if uploaded_photos and not st.session_state.restaurant_results:
        st.caption(f"已上传 {len(uploaded_photos)} 张照片" +
                   (f" + 1 个已有 Excel" if existing_excel else ""))

        if st.button("🚀 开始识别", type="primary"):
            if existing_excel:
                try:
                    st.session_state.restaurant_existing_wb = load_workbook(existing_excel)
                except Exception as e:
                    st.error(f"读取已有 Excel 失败: {e}")
                    st.session_state.restaurant_existing_wb = None
            else:
                st.session_state.restaurant_existing_wb = None

            client = OpenAI(api_key=api_key,
                            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1")
            all_days = []
            prog = st.progress(0, text="正在识别中...")
            total = len(uploaded_photos)
            for idx, f in enumerate(uploaded_photos):
                prog.progress(idx / total, text=f"正在识别第 {idx+1}/{total} 张: {f.name}...")
                image = Image.open(f)
                try:
                    raw = _call_vl(client, _img_b64(image), REPORT_PROMPT)
                    day = _parse_json(raw)
                    day["_filename"] = f.name
                    all_days.append(day)
                except Exception as e:
                    st.error(f"识别 {f.name} 失败: {e}")
            prog.progress(1.0, text="全部识别完成！")

            # 检查哪些日期在已有 Excel 中已有实际流水
            if st.session_state.restaurant_existing_wb:
                wb_exist = st.session_state.restaurant_existing_wb
                skipped = []
                for day in all_days:
                    try:
                        dt = datetime.strptime(day["date"], "%Y-%m-%d")
                        sn = str(dt.day)
                        if sn in wb_exist.sheetnames and _sheet_has_data(wb_exist[sn]):
                            skipped.append(day["date"])
                    except Exception:
                        pass
                if skipped:
                    st.warning(f"以下日期在已有 Excel 中已有流水数据，将跳过不覆盖：{', '.join(skipped)}")

            for day in all_days:
                _process_notes(day)
                _validate_day(day)

            st.session_state.restaurant_results = all_days
            st.session_state.restaurant_confirmed = False
            st.rerun()

    # --- 确认编辑 ---
    if st.session_state.restaurant_results and not st.session_state.restaurant_confirmed:
        all_days = st.session_state.restaurant_results

        for day in all_days:
            _validate_day(day)

        total_warnings = sum(
            len(r.get("_warnings", [])) for d in all_days for r in d.get("rows", []))

        if total_warnings:
            st.warning(f"共识别 **{len(all_days)}** 天数据，"
                       f"发现 **{total_warnings}** 处校验异常（🔴），请重点检查：")
        else:
            st.success(f"共识别 **{len(all_days)}** 天的数据，校验全部通过，请确认：")

        for di, day in enumerate(all_days):
            date_str = day.get("date", "未知")
            rows = day.get("rows", [])
            unc = _has_uncertain(day)
            icon = "⚠️" if unc else "✅"

            with st.expander(f"{icon} {date_str}（{day.get('_filename','')}）- {len(rows)} 条记录",
                             expanded=unc):
                all_days[di]["date"] = st.text_input(
                    "日期 (YYYY-MM-DD)", value=date_str, key=f"dt_{di}")

                for ri, row in enumerate(rows):
                    uf = []
                    for k in row:
                        if k.endswith("_uncertain") and row[k]:
                            uf.append(k.replace("_uncertain", ""))
                    for p in row.get("payments", []):
                        if p.get("uncertain"):
                            uf.append(f"付款({p.get('method','?')})")

                    label = f"**第 {row.get('seq','?')} 行**"
                    if uf:
                        label += f" ⚠️ 不确定: {', '.join(uf)}"

                    row_warnings = row.get("_warnings", [])
                    if row_warnings:
                        label += "  🔴 " + "；".join(row_warnings)

                    st.markdown(label)

                    c1, c2, c3, c4 = st.columns(4)
                    with c1:
                        row["period"] = st.text_input(
                            "餐段(中/晚)", value=row.get("period", ""),
                            key=f"pd_{di}_{ri}")
                    with c2:
                        row["room"] = st.text_input(
                            "包间号", value=row.get("room", ""),
                            key=f"rm_{di}_{ri}")
                    with c3:
                        row["revenue"] = st.number_input(
                            "营业额", value=float(row.get("revenue") or 0),
                            step=1.0, format="%.0f", key=f"rv_{di}_{ri}")
                    with c4:
                        row["income"] = st.number_input(
                            "收入", value=float(row.get("income") or 0),
                            step=1.0, format="%.0f", key=f"ic_{di}_{ri}")

                    r1, r2 = st.columns(2)
                    with r1:
                        row["drinks"] = st.number_input(
                            "酒水", value=float(row.get("drinks") or 0),
                            step=1.0, format="%.0f", key=f"dk_{di}_{ri}")
                    with r2:
                        row["row_note"] = st.text_input(
                            "行备注", value=row.get("row_note", ""),
                            key=f"rn_{di}_{ri}")

                    payments = row.get("payments", [{"method": "", "amount": 0}])
                    pay_sum = sum(p.get("amount") or 0 for p in payments)
                    income_val = row.get("income") or 0
                    pay_label = f"💳 支付明细（合计 {pay_sum}）"
                    if abs(pay_sum - income_val) > 0.5:
                        pay_label += f" 🔴 ≠ 收入({income_val})"
                    st.markdown(pay_label)

                    to_delete = None
                    new_payments = []
                    for pi, pay in enumerate(payments):
                        pc1, pc2, pc3 = st.columns([4, 4, 1])
                        with pc1:
                            m = st.text_input(
                                f"方式{pi+1}", value=pay.get("method", ""),
                                key=f"pm_{di}_{ri}_{pi}",
                                help="微信/支付宝/现金/抖音/会员卡/挂帐/饿了么/美团/收钱吧")
                        with pc2:
                            a = st.number_input(
                                f"金额{pi+1}", value=float(pay.get("amount") or 0),
                                step=1.0, format="%.0f", key=f"pa_{di}_{ri}_{pi}")
                        with pc3:
                            st.markdown("<br>", unsafe_allow_html=True)
                            if st.button("🗑️", key=f"del_{di}_{ri}_{pi}",
                                         help="删除此支付方式"):
                                to_delete = pi
                        new_payments.append({"method": m, "amount": a})

                    if to_delete is not None and len(new_payments) > 1:
                        new_payments.pop(to_delete)
                        row["payments"] = new_payments
                        st.rerun()
                    row["payments"] = new_payments

                    if st.button(f"➕ 添加支付方式", key=f"add_{di}_{ri}"):
                        row["payments"].append({"method": "", "amount": 0})
                        st.rerun()

                    st.markdown("---")

                all_days[di]["notes"] = st.text_area(
                    "备注", value=day.get("notes", ""), key=f"nt_{di}", height=60)

        if st.button("✅ 确认无误，生成报表", type="primary"):
            for day in all_days:
                _validate_day(day)
            remaining = sum(
                len(r.get("_warnings", [])) for d in all_days for r in d.get("rows", []))
            if remaining:
                st.error(f"仍有 {remaining} 处校验未通过（付款合计≠收入 或 收入>营业额），"
                         "请修正后再确认。如确认数据无误，请再次点击按钮强制生成。")
                st.session_state["_force_gen"] = True
            else:
                st.session_state.restaurant_results = all_days
                st.session_state.restaurant_confirmed = True
                st.rerun()

        if st.session_state.get("_force_gen"):
            if st.button("⚠️ 忽略校验异常，强制生成报表"):
                st.session_state.restaurant_results = all_days
                st.session_state.restaurant_confirmed = True
                st.session_state["_force_gen"] = False
                st.rerun()

    # --- 生成 & 下载 ---
    if st.session_state.restaurant_confirmed and st.session_state.restaurant_results:
        try:
            excel = _build_excel(
                st.session_state.restaurant_results,
                existing_wb=st.session_state.restaurant_existing_wb,
            )
        except Exception as e:
            st.error(f"❌ 生成报表时出错：{e}")
            st.exception(e)
        else:
            st.success("✅ 报表生成完成！如模板含合并单元格，请打开 Excel 查看末尾的"
                       "『_诊断说明』工作表确认处理情况。")
            st.download_button(
                "📥 下载餐厅日报表", data=excel,
                file_name=f"餐厅日报表_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary")
        if st.button("🔄 重新开始"):
            st.session_state.restaurant_results = None
            st.session_state.restaurant_confirmed = False
            st.session_state.restaurant_existing_wb = None
            st.rerun()
