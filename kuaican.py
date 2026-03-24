import streamlit as st
import base64
import json
import re
import io
import math
from datetime import datetime
from collections import defaultdict
from PIL import Image
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MEAL_PRICE = 15
SLICE_HEIGHT = 3000
OVERLAP = 200

EXTRACTION_PROMPT = """你是一个收款记录 OCR 助手。请仔细识别这张收款截图中的**所有**收款记录，并以严格的 JSON 数组格式输出。

每条记录包含以下字段：
- "amount": 收款金额（数字，保留两位小数）
- "time": 支付时间（格式 "YYYY-MM-DD HH:MM:SS"）
- "method": 支付方式（只能是 "微信" 或 "支付宝"）

判断支付方式的依据：
- 绿色圆形勾号图标 = 微信
- 蓝色圆形图标（带"支"字或支付宝标志）= 支付宝

注意事项：
1. 必须提取截图中的每一条记录，不要遗漏
2. 金额请提取实际数字，如 15.00、16.00、30.00 等
3. 只输出 JSON 数组，不要输出任何其他文字
4. 如果某条记录看不清，尽量识别，不要跳过

输出示例：
[
  {"amount": 15.00, "time": "2026-03-24 11:44:40", "method": "微信"},
  {"amount": 16.00, "time": "2026-03-24 11:44:13", "method": "支付宝"}
]"""


def _split_long_image(image):
    w, h = image.size
    if h <= SLICE_HEIGHT + OVERLAP:
        return [image]
    slices, y = [], 0
    while y < h:
        bottom = min(y + SLICE_HEIGHT, h)
        slices.append(image.crop((0, y, w, bottom)))
        y += SLICE_HEIGHT - OVERLAP
        if bottom == h:
            break
    return slices


def _img_b64(image):
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def _call_vl(client, b64, prompt):
    c = client.chat.completions.create(
        model="qwen-vl-max",
        messages=[{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
            {"type": "text", "text": prompt},
        ]}],
    )
    return c.choices[0].message.content


def _parse_json(raw):
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```\w*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"[\[{].*[}\]]", text, re.DOTALL)
        if m:
            return json.loads(m.group())
        raise ValueError(f"无法解析:\n{raw[:500]}")


def _dedup(records):
    seen, out = set(), []
    for r in records:
        k = (r["amount"], r["time"])
        if k not in seen:
            seen.add(k)
            out.append(r)
    return out


def _classify(amount):
    n = max(1, math.floor(amount / MEAL_PRICE))
    rem = round(amount - MEAL_PRICE * n, 2)
    if rem <= 0:
        return {"meal": amount, "box": 0, "drink": 0, "note": "仅餐费"}
    if rem <= n:
        return {"meal": MEAL_PRICE * n, "box": rem, "drink": 0, "note": "餐费+打包盒"}
    return {"meal": MEAL_PRICE * n, "box": 0, "drink": rem, "note": "餐费+饮料"}


def _period(ts):
    try:
        return "午餐" if int(ts[11:13]) < 15 else "晚餐"
    except Exception:
        return "未知"


def _build_excel(records):
    wb = Workbook()
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, size=11, color="FFFFFF")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")
    mfmt = '#,##0.00'

    def wh(ws, headers, widths):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font, c.fill, c.alignment, c.border = hfont, hfill, center, border
            ws.column_dimensions[c.column_letter].width = w

    def wr(ws, row, vals, mc=None):
        mc = mc or set()
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.border, c.alignment = border, center
            if i in mc:
                c.number_format = mfmt

    ws1 = wb.active
    ws1.title = "收款明细"
    wh(ws1, ["序号", "日期", "金额", "支付时间", "餐段", "支付方式", "餐费", "打包盒费", "饮料费", "备注"],
       [6, 12, 10, 22, 8, 10, 10, 10, 10, 14])

    daily = defaultdict(lambda: dict(income=0, count=0, meal=0, box=0, drink=0,
                                     wechat=0, wechat_n=0, alipay=0, alipay_n=0))
    tm = tb = td = ti = wt = at = 0
    wc = ac = 0
    for i, r in enumerate(records, 1):
        a = r["amount"]
        cl = _classify(a)
        ds = r["time"][:10]
        wr(ws1, i+1, [i, ds, a, r["time"], _period(r["time"]), r["method"],
                       cl["meal"], cl["box"], cl["drink"], cl["note"]], {3,7,8,9})
        ti += a; tm += cl["meal"]; tb += cl["box"]; td += cl["drink"]
        is_wx = r["method"] == "微信"
        if is_wx: wt += a; wc += 1
        else: at += a; ac += 1
        d = daily[ds]
        d["income"] += a; d["count"] += 1; d["meal"] += cl["meal"]
        d["box"] += cl["box"]; d["drink"] += cl["drink"]
        if is_wx: d["wechat"] += a; d["wechat_n"] += 1
        else: d["alipay"] += a; d["alipay_n"] += 1

    ws2 = wb.create_sheet("日汇总")
    wh(ws2, ["日期","总收入","总笔数","就餐人数","总餐费","总打包盒费","总饮料费",
             "微信收款","微信笔数","支付宝收款","支付宝笔数"],
       [12,12,8,10,12,12,12,12,10,12,10])
    for ri, ds in enumerate(sorted(daily), 2):
        d = daily[ds]
        wr(ws2, ri, [ds, round(d["income"],2), d["count"], int(round(d["meal"]/MEAL_PRICE)),
                      round(d["meal"],2), round(d["box"],2), round(d["drink"],2),
                      round(d["wechat"],2), d["wechat_n"], round(d["alipay"],2), d["alipay_n"]],
           {2,5,6,7,8,10})

    ws3 = wb.create_sheet("月汇总")
    dates = sorted(daily)
    dr = dates[0] if len(dates)==1 else f"{dates[0]} ~ {dates[-1]}" if dates else "未知"
    rows = [("项目","金额/数量"),("日期范围",dr),("总营业收入",round(ti,2)),
            ("总笔数",len(records)),("就餐人数",int(round(tm/MEAL_PRICE))),
            ("总餐费",round(tm,2)),("总打包盒费",round(tb,2)),("总饮料费",round(td,2)),
            ("",""),("微信收款总额",round(wt,2)),("微信笔数",wc),
            ("支付宝收款总额",round(at,2)),("支付宝笔数",ac)]
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 16
    for ri, (l, v) in enumerate(rows, 1):
        c1 = ws3.cell(row=ri, column=1, value=l)
        c2 = ws3.cell(row=ri, column=2, value=v)
        for c in (c1, c2):
            c.border, c.alignment = border, center
        if ri == 1:
            for c in (c1, c2): c.font, c.fill = hfont, hfill
        elif isinstance(v, float):
            c2.number_format = mfmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ========== UI 入口 ==========

def run():
    st.title("🍚 十五元快餐 · 收款分析")
    st.caption("上传收款截图，自动识别每笔订单并生成 Excel 报表")

    if st.button("⬅️ 返回主页"):
        st.session_state.current_page = "home"
        st.rerun()

    api_key = st.secrets.get("DASHSCOPE_API_KEY", "") if hasattr(st, "secrets") else ""
    if not api_key:
        api_key = st.text_input("请输入阿里云百炼 API Key（sk- 开头）", type="password")
    if not api_key:
        st.info("请先配置 API Key 才能使用。")
        st.stop()

    uploaded = st.file_uploader("上传收款截图", type=["png","jpg","jpeg"],
                                help="支持微信/支付宝收款助手的长截图")
    if not uploaded:
        return

    image = Image.open(uploaded)
    st.caption(f"已上传：{uploaded.name}（{image.size[0]}×{image.size[1]}px）")
    run_clicked = st.button("🚀 开始分析", type="primary")
    with st.expander("查看上传的截图", expanded=False):
        st.image(image, width="stretch")

    if not run_clicked:
        return

    client = OpenAI(api_key=api_key, base_url="https://dashscope.aliyuncs.com/compatible-mode/v1")
    slices = _split_long_image(image)
    total = len(slices)
    prog = st.progress(0, text="正在识别中...")
    all_rec = []
    for idx, sl in enumerate(slices):
        prog.progress(idx/total, text=f"正在识别第 {idx+1}/{total} 段...")
        try:
            raw = _call_vl(client, _img_b64(sl), EXTRACTION_PROMPT)
            all_rec.extend(_parse_json(raw))
        except Exception as e:
            st.error(f"第 {idx+1} 段识别失败: {e}")
    prog.progress(1.0, text="识别完成！")

    all_rec = _dedup(all_rec)
    try: all_rec.sort(key=lambda r: r["time"])
    except Exception: pass

    if not all_rec:
        st.error("未能识别到任何收款记录。")
        return

    st.success(f"共识别 **{len(all_rec)}** 笔收款记录")
    md = ["| 序号 | 金额 | 支付时间 | 支付方式 | 餐费 | 打包盒费 | 饮料费 | 备注 |",
          "|:---:|-----:|:-------:|:------:|-----:|-------:|------:|:----:|"]
    for i, r in enumerate(all_rec, 1):
        c = _classify(r["amount"])
        md.append(f"| {i} | {r['amount']:.2f} | {r['time']} | {r['method']} "
                  f"| {c['meal']:.2f} | {c['box']:.2f} | {c['drink']:.2f} | {c['note']} |")
    st.markdown("\n".join(md))

    ti = sum(r["amount"] for r in all_rec)
    tm = sum(_classify(r["amount"])["meal"] for r in all_rec)
    tb = sum(_classify(r["amount"])["box"] for r in all_rec)
    td = sum(_classify(r["amount"])["drink"] for r in all_rec)
    st.markdown("---")
    st.markdown(f"**总收入** ¥{ti:.2f} &nbsp;|&nbsp; **餐费** ¥{tm:.2f} &nbsp;|&nbsp; "
                f"**打包盒费** ¥{tb:.2f} &nbsp;|&nbsp; **饮料费** ¥{td:.2f}")

    excel = _build_excel(all_rec)
    st.download_button("📥 下载 Excel 报表", data=excel,
                       file_name=f"收款分析_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary")
