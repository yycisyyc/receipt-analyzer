import streamlit as st
import base64
import json
import re
import io
import math
from datetime import datetime
from PIL import Image
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# 页面配置
# ---------------------------------------------------------------------------
st.set_page_config(page_title="十五元快餐 - 收款分析", page_icon="🍚", layout="centered")

st.title("🍚 十五元快餐 · 收款分析工具")
st.caption("上传收款截图，自动识别每笔订单并生成 Excel 报表")

# ---------------------------------------------------------------------------
# API Key（优先从 Streamlit Secrets 读取，也支持手动输入）
# ---------------------------------------------------------------------------
api_key = st.secrets.get("DASHSCOPE_API_KEY", "") if hasattr(st, "secrets") else ""

if not api_key:
    api_key = st.text_input(
        "请输入阿里云百炼 API Key（sk- 开头）",
        type="password",
        help="在 https://bailian.console.aliyun.com 获取",
    )

if not api_key:
    st.info("请先配置 API Key 才能使用。")
    st.stop()

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
MEAL_PRICE = 15
SLICE_HEIGHT = 3000
OVERLAP = 200

EXTRACTION_PROMPT = """你是一个收款记录 OCR 助手。请仔细识别这张收款截图中的**所有**收款记录，并以严格的 JSON 数组格式输出。

每条记录包含以下字段：
- "amount": 收款金额（数字，保留两位小数）
- "time": 支付时间（格式 "YYYY-MM-DD HH:MM:SS"）
- "method": 支付方式（只能是 "微信" 或 "支付宝"）

判断支付方式的依据：
- 绿色圆形勾号图标 = 微信
- 蓝色圆形图标（带"支"字或支付宝标志）= 支付宝

注意事项：
1. 必须提取截图中的每一条记录，不要遗漏
2. 金额请提取实际数字，如 15.00、16.00、30.00 等
3. 只输出 JSON 数组，不要输出任何其他文字
4. 如果某条记录看不清，尽量识别，不要跳过

输出示例：
[
  {"amount": 15.00, "time": "2026-03-24 11:44:40", "method": "微信"},
  {"amount": 16.00, "time": "2026-03-24 11:44:13", "method": "支付宝"}
]"""

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def split_long_image(image: Image.Image) -> list[Image.Image]:
    """将长截图按固定高度切分，相邻切片保留重叠区域。"""
    w, h = image.size
    if h <= SLICE_HEIGHT + OVERLAP:
        return [image]

    slices = []
    y = 0
    while y < h:
        bottom = min(y + SLICE_HEIGHT, h)
        slices.append(image.crop((0, y, w, bottom)))
        y += SLICE_HEIGHT - OVERLAP
        if bottom == h:
            break
    return slices


def image_to_base64(image: Image.Image) -> str:
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def call_qwen_vl(client: OpenAI, b64: str) -> str:
    completion = client.chat.completions.create(
        model="qwen-vl-max",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{b64}"},
                    },
                    {"type": "text", "text": EXTRACTION_PROMPT},
                ],
            }
        ],
    )
    return completion.choices[0].message.content


def parse_records(raw: str) -> list[dict]:
    """从模型返回的文本中提取 JSON 数组。"""
    # 尝试直接解析
    text = raw.strip()
    # 去除可能的 markdown 代码块标记
    if text.startswith("```"):
        text = re.sub(r"^```\w*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        text = text.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # 尝试找到 JSON 数组部分
        match = re.search(r"\[.*\]", text, re.DOTALL)
        if match:
            return json.loads(match.group())
        raise ValueError(f"无法解析模型返回内容:\n{raw[:500]}")


def deduplicate(records: list[dict]) -> list[dict]:
    """按 (amount, time) 去重，保留首次出现的记录。"""
    seen = set()
    result = []
    for r in records:
        key = (r["amount"], r["time"])
        if key not in seen:
            seen.add(key)
            result.append(r)
    return result


def classify_amount(amount: float) -> dict:
    """根据业务规则将金额拆分为 餐费 / 打包盒费 / 饮料费。"""
    n = max(1, math.floor(amount / MEAL_PRICE))
    remainder = round(amount - MEAL_PRICE * n, 2)

    if remainder < 0:
        # 金额不足一份餐（如 10 元），全部算餐费
        return {"meal": amount, "box": 0, "drink": 0, "note": "仅餐费"}

    if remainder == 0:
        return {"meal": amount, "box": 0, "drink": 0, "note": "仅餐费"}
    elif remainder <= n:
        return {
            "meal": MEAL_PRICE * n,
            "box": remainder,
            "drink": 0,
            "note": "餐费+打包盒",
        }
    else:
        return {
            "meal": MEAL_PRICE * n,
            "box": 0,
            "drink": remainder,
            "note": "餐费+饮料",
        }


def build_excel(records: list[dict]) -> bytes:
    """生成 Excel 文件并返回二进制内容。"""
    wb = Workbook()

    # ---- 样式 ----
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    money_fmt = '#,##0.00'

    # ========== Sheet 1: 收款明细 ==========
    ws1 = wb.active
    ws1.title = "收款明细"
    headers = ["序号", "金额", "支付时间", "支付方式", "餐费", "打包盒费", "饮料费", "备注"]
    col_widths = [6, 10, 22, 10, 10, 10, 10, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws1.column_dimensions[cell.column_letter].width = w

    total_meal = total_box = total_drink = total_income = 0
    wechat_total = alipay_total = 0
    wechat_count = alipay_count = 0

    for i, rec in enumerate(records, 1):
        amt = rec["amount"]
        cls = classify_amount(amt)
        row_data = [
            i,
            amt,
            rec["time"],
            rec["method"],
            cls["meal"],
            cls["box"],
            cls["drink"],
            cls["note"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i + 1, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center
            if col_idx in (2, 5, 6, 7):
                cell.number_format = money_fmt

        total_income += amt
        total_meal += cls["meal"]
        total_box += cls["box"]
        total_drink += cls["drink"]
        if rec["method"] == "微信":
            wechat_total += amt
            wechat_count += 1
        else:
            alipay_total += amt
            alipay_count += 1

    # ========== Sheet 2: 汇总 ==========
    ws2 = wb.create_sheet("汇总")
    summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # 从记录中提取日期范围
    dates = []
    for rec in records:
        try:
            dates.append(rec["time"][:10])
        except Exception:
            pass
    if dates:
        date_min, date_max = min(dates), max(dates)
        date_range = date_min if date_min == date_max else f"{date_min} ~ {date_max}"
    else:
        date_range = "未知"

    summary = [
        ("项目", "金额/数量"),
        ("日期范围", date_range),
        ("总营业收入", round(total_income, 2)),
        ("总笔数", len(records)),
        ("总餐费", round(total_meal, 2)),
        ("总打包盒费", round(total_box, 2)),
        ("总饮料费", round(total_drink, 2)),
        ("", ""),
        ("微信收款总额", round(wechat_total, 2)),
        ("微信笔数", wechat_count),
        ("支付宝收款总额", round(alipay_total, 2)),
        ("支付宝笔数", alipay_count),
    ]

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 16

    for row_idx, (label, value) in enumerate(summary, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c2 = ws2.cell(row=row_idx, column=2, value=value)
        c1.border = thin_border
        c2.border = thin_border
        c1.alignment = center
        c2.alignment = center
        if row_idx == 1:
            c1.font = header_font_white
            c2.font = header_font_white
            c1.fill = summary_header_fill
            c2.fill = summary_header_fill
        elif isinstance(value, float):
            c2.number_format = money_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

uploaded = st.file_uploader(
    "上传收款截图",
    type=["png", "jpg", "jpeg"],
    help="支持微信/支付宝收款助手的长截图",
)

if uploaded:
    image = Image.open(uploaded)

    # 先显示文件信息和按钮，图片折叠在下面
    st.caption(f"已上传：{uploaded.name}（{image.size[0]}×{image.size[1]}px）")
    run_clicked = st.button("🚀 开始分析", type="primary")

    with st.expander("查看上传的截图", expanded=False):
        st.image(image, width="stretch")

    if run_clicked:
        client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )

        slices = split_long_image(image)
        total_slices = len(slices)

        progress = st.progress(0, text="正在识别中...")
        all_records = []

        for idx, sl in enumerate(slices):
            progress.progress(
                (idx) / total_slices,
                text=f"正在识别第 {idx + 1}/{total_slices} 段...",
            )
            b64 = image_to_base64(sl)
            try:
                raw = call_qwen_vl(client, b64)
                records = parse_records(raw)
                all_records.extend(records)
            except Exception as e:
                st.error(f"第 {idx + 1} 段识别失败: {e}")

        progress.progress(1.0, text="识别完成！正在生成报表...")

        # 去重 & 按时间排序
        all_records = deduplicate(all_records)
        try:
            all_records.sort(key=lambda r: r["time"])
        except Exception:
            pass

        if not all_records:
            st.error("未能识别到任何收款记录，请检查截图是否正确。")
            st.stop()

        # 展示明细
        st.success(f"共识别 **{len(all_records)}** 笔收款记录")

        display_data = []
        for i, rec in enumerate(all_records, 1):
            cls = classify_amount(rec["amount"])
            display_data.append({
                "序号": i,
                "金额": rec["amount"],
                "支付时间": rec["time"],
                "支付方式": rec["method"],
                "餐费": cls["meal"],
                "打包盒费": cls["box"],
                "饮料费": cls["drink"],
                "备注": cls["note"],
            })

        md_lines = [
            "| 序号 | 金额 | 支付时间 | 支付方式 | 餐费 | 打包盒费 | 饮料费 | 备注 |",
            "|:---:|-----:|:-------:|:------:|-----:|-------:|------:|:----:|",
        ]
        for row in display_data:
            md_lines.append(
                f"| {row['序号']} | {row['金额']:.2f} | {row['支付时间']} | {row['支付方式']} "
                f"| {row['餐费']:.2f} | {row['打包盒费']:.2f} | {row['饮料费']:.2f} | {row['备注']} |"
            )
        st.markdown("\n".join(md_lines))

        # 汇总
        total_income = sum(r["amount"] for r in all_records)
        total_meal = sum(classify_amount(r["amount"])["meal"] for r in all_records)
        total_box = sum(classify_amount(r["amount"])["box"] for r in all_records)
        total_drink = sum(classify_amount(r["amount"])["drink"] for r in all_records)

        st.markdown("---")
        st.markdown(
            f"**总收入** ¥{total_income:.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"**餐费** ¥{total_meal:.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"**打包盒费** ¥{total_box:.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"**饮料费** ¥{total_drink:.2f}"
        )

        # 生成 Excel 并提供下载
        excel_bytes = build_excel(all_records)
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")

        st.download_button(
            label="📥 下载 Excel 报表",
            data=excel_bytes,
            file_name=f"收款分析_{now_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
