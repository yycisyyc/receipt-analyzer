import streamlit as st
import base64
import json
import re
import io
from datetime import datetime
from PIL import Image
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.title("📋 餐厅日报表")
st.caption("上传手写日报照片，AI 识别后生成标准 Excel 日报表")

# ---------------------------------------------------------------------------
# API Key
# ---------------------------------------------------------------------------
api_key = st.secrets.get("DASHSCOPE_API_KEY", "") if hasattr(st, "secrets") else ""
if not api_key:
    api_key = st.text_input(
        "请输入阿里云百炼 API Key（sk- 开头）",
        type="password",
        help="在 https://bailian.console.aliyun.com 获取",
    )
if not api_key:
    st.info("请先配置 API Key 才能使用。")
    st.stop()

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
VALID_ROOMS = (
    [f"卡{i}" for i in range(1, 7)]
    + ["111", "222", "333", "555", "666", "777", "888", "999"]
    + [f"厅{i}" for i in range(1, 21)]
)

VALID_METHODS = ["支付宝", "抖音", "微信", "现金", "饿了么", "美团", "收钱吧"]

REPORT_PROMPT = """你是一个手写餐厅日报 OCR 助手。请仔细识别这张手写日报照片中的所有信息。

**照片结构说明**：
- 右上角有日期（年/月/日）
- 表格从左到右的列依次是：序号、用餐时间(中/晚)、包间号、营业额、(空列)、收入、付款方式、实际收款
- 表格下方可能有备注信息

**包间号只可能是以下值之一**：
卡1, 卡2, 卡3, 卡4, 卡5, 卡6, 111, 222, 333, 555, 666, 777, 888, 999, 厅1~厅20

**付款方式只可能是以下值（可多选，用"/"分隔）**：
支付宝, 抖音, 微信, 现金, 饿了么, 美团, 收钱吧

请以严格的 JSON 格式输出，结构如下：
{
  "date": "YYYY-MM-DD",
  "rows": [
    {
      "seq": 1,
      "period": "中或晚",
      "period_uncertain": false,
      "room": "包间号",
      "room_uncertain": false,
      "revenue": 数字(营业额),
      "revenue_uncertain": false,
      "income": 数字(收入),
      "income_uncertain": false,
      "payment": "付款方式",
      "payment_uncertain": false,
      "actual": 数字(实际收款),
      "actual_uncertain": false
    }
  ],
  "notes": "底部备注内容，没有则为空字符串"
}

**重要规则**：
1. 只提取有实际数据的行，空行跳过
2. 对于你不确定的字段，将对应的 xxx_uncertain 设为 true
3. 如果某个字段完全看不清，填 null 并标记 uncertain 为 true
4. 日期从右上角识别，格式为 YYYY-MM-DD
5. 营业额和收入是数字，不要带逗号或其他符号
6. 只输出 JSON，不要输出任何其他文字"""

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def image_to_base64(image: Image.Image) -> str:
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def call_qwen_vl(client: OpenAI, b64: str, prompt: str) -> str:
    completion = client.chat.completions.create(
        model="qwen-vl-max",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
    )
    return completion.choices[0].message.content


def parse_json_obj(raw: str) -> dict:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```\w*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if match:
            return json.loads(match.group())
        raise ValueError(f"无法解析模型返回内容:\n{raw[:500]}")


def has_uncertain(day_data: dict) -> bool:
    for row in day_data.get("rows", []):
        for key in row:
            if key.endswith("_uncertain") and row[key]:
                return True
    return False


def build_restaurant_excel(all_days: list[dict]) -> bytes:
    """按模板格式生成餐厅日报 Excel。"""
    wb = Workbook()

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    money_fmt = '#,##0.00'

    HEADERS = ["序号", "中餐/晚餐", "包间号", "营业额", "折扣", "收入",
               "手续费", "充值", "实收", "挂账", "挂账收回",
               "会员卡赠送", "会员卡消费", "会员卡余额", "付款方式", "酒水", "备注"]
    COL_WIDTHS = [6, 10, 8, 10, 8, 10, 10, 8, 10, 8, 10, 10, 10, 10, 12, 8, 14]

    # 按日期排序
    all_days_sorted = sorted(all_days, key=lambda d: d.get("date", ""))

    # 确定月份信息，用于 sheet 命名
    if all_days_sorted:
        first_date = all_days_sorted[0]["date"]
        try:
            base_dt = datetime.strptime(first_date, "%Y-%m-%d")
            year, month = base_dt.year, base_dt.month
        except Exception:
            year, month = 2026, 1
    else:
        year, month = 2026, 1

    first_sheet = True
    for day_data in all_days_sorted:
        date_str = day_data.get("date", "")
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            day_num = dt.day
        except Exception:
            day_num = 0

        sheet_name = str(day_num) if day_num > 0 else date_str

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        # Row 1: date header in K1
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            ws["K1"] = f"日期：  {dt.year}  年  {dt.month}  月  {dt.day}  日"
        except Exception:
            ws["K1"] = f"日期：{date_str}"
        ws["K1"].font = bold_font

        # Row 2: column headers
        for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = w

        # Rows 3-27: data rows (sequence numbers 1-25)
        rows = day_data.get("rows", [])
        for row_idx in range(3, 28):
            seq = row_idx - 2
            ws.cell(row=row_idx, column=1, value=seq).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = center
            # Fill empty borders for all columns
            for c in range(2, 18):
                ws.cell(row=row_idx, column=c).border = thin_border
                ws.cell(row=row_idx, column=c).alignment = center

        for row in rows:
            seq = row.get("seq", 0)
            if seq < 1 or seq > 25:
                continue
            r = seq + 2

            period = row.get("period", "")
            room = row.get("room", "")
            revenue = row.get("revenue")
            income = row.get("income")
            payment = row.get("payment", "")
            actual = row.get("actual")

            ws.cell(row=r, column=2, value=period)
            ws.cell(row=r, column=3, value=room)

            if revenue is not None:
                ws.cell(row=r, column=4, value=revenue)
                ws.cell(row=r, column=4).number_format = money_fmt

            # 折扣(E) = 营业额(D) - 收入(F)
            if revenue is not None and income is not None:
                discount = round(revenue - income, 2)
                if discount > 0:
                    ws.cell(row=r, column=5, value=discount)
                    ws.cell(row=r, column=5).number_format = money_fmt

            if income is not None:
                ws.cell(row=r, column=6, value=income)
                ws.cell(row=r, column=6).number_format = money_fmt
                # 手续费(G) = 收入 * 0.0038
                ws.cell(row=r, column=7).value = f"=F{r}*0.0038"
                ws.cell(row=r, column=7).number_format = money_fmt
                # 实收(I) = 收入 - 手续费
                ws.cell(row=r, column=9).value = f"=F{r}-G{r}"
                ws.cell(row=r, column=9).number_format = money_fmt

            ws.cell(row=r, column=15, value=payment)

            if actual is not None:
                ws.cell(row=r, column=9, value=actual)
                ws.cell(row=r, column=9).number_format = money_fmt

        # Row 28: 合计
        ws.cell(row=28, column=1, value="合计").font = bold_font
        ws.cell(row=28, column=1).border = thin_border
        ws.cell(row=28, column=1).alignment = center
        for col_idx in range(4, 18):
            col_letter = ws.cell(row=2, column=col_idx).column_letter
            cell = ws.cell(row=28, column=col_idx)
            cell.value = f"=SUM({col_letter}3:{col_letter}27)"
            cell.number_format = money_fmt
            cell.border = thin_border
            cell.alignment = center
            cell.font = bold_font

        # Rows 29-38: revenue/expenditure breakdown section
        ws.cell(row=29, column=1, value="营业收入：").font = bold_font
        ws.cell(row=29, column=6, value="采购支出：").font = bold_font
        ws.cell(row=30, column=1, value="收钱吧")
        ws.cell(row=30, column=2, value="微信/支付宝")
        ws.cell(row=30, column=6, value="食材（肉，水产，蔬菜，调味品，饮料）")
        ws.cell(row=32, column=2, value="现金")
        ws.cell(row=33, column=2, value="抖音/团券")
        ws.cell(row=33, column=6, value="其它费用")
        ws.cell(row=34, column=2, value="会员卡")
        ws.cell(row=34, column=6, value="设备、维修")
        ws.cell(row=35, column=6, value="水电燃气费")
        ws.cell(row=36, column=6, value="工资")
        ws.cell(row=37, column=6, value="营销")
        ws.cell(row=38, column=1, value="合计").font = bold_font
        ws.cell(row=38, column=3).value = "=SUM(C30:C36)"

        # 备注
        notes = day_data.get("notes", "")
        if notes:
            ws.cell(row=29, column=10, value=f"备注：{notes}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Session State 管理
# ---------------------------------------------------------------------------

if "restaurant_results" not in st.session_state:
    st.session_state.restaurant_results = None
if "restaurant_confirmed" not in st.session_state:
    st.session_state.restaurant_confirmed = False

# ---------------------------------------------------------------------------
# Step 1: 上传照片
# ---------------------------------------------------------------------------

uploaded_files = st.file_uploader(
    "上传手写日报照片（可多选）",
    type=["png", "jpg", "jpeg"],
    accept_multiple_files=True,
    help="每张照片对应一天的手写日报",
)

if uploaded_files and not st.session_state.restaurant_results:
    st.caption(f"已上传 {len(uploaded_files)} 张照片")

    if st.button("🚀 开始识别", type="primary"):
        client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )

        all_days = []
        progress = st.progress(0, text="正在识别中...")
        total = len(uploaded_files)

        for idx, f in enumerate(uploaded_files):
            progress.progress(idx / total, text=f"正在识别第 {idx + 1}/{total} 张: {f.name}...")
            image = Image.open(f)
            b64 = image_to_base64(image)
            try:
                raw = call_qwen_vl(client, b64, REPORT_PROMPT)
                day_data = parse_json_obj(raw)
                day_data["_filename"] = f.name
                all_days.append(day_data)
            except Exception as e:
                st.error(f"识别 {f.name} 失败: {e}")

        progress.progress(1.0, text="全部识别完成！")
        st.session_state.restaurant_results = all_days
        st.session_state.restaurant_confirmed = False
        st.rerun()

# ---------------------------------------------------------------------------
# Step 2 & 3: 展示结果 & 用户确认
# ---------------------------------------------------------------------------

if st.session_state.restaurant_results and not st.session_state.restaurant_confirmed:
    all_days = st.session_state.restaurant_results

    st.success(f"共识别 **{len(all_days)}** 天的数据，请检查并修改不准确的内容：")

    has_any_uncertain = False

    for day_idx, day_data in enumerate(all_days):
        date_str = day_data.get("date", "未知日期")
        filename = day_data.get("_filename", "")
        rows = day_data.get("rows", [])
        notes = day_data.get("notes", "")
        day_uncertain = has_uncertain(day_data)
        if day_uncertain:
            has_any_uncertain = True

        icon = "⚠️" if day_uncertain else "✅"
        with st.expander(f"{icon} {date_str}（{filename}）- {len(rows)} 条记录", expanded=day_uncertain):
            # 日期修改
            new_date = st.text_input(
                "日期", value=date_str, key=f"date_{day_idx}",
                help="格式 YYYY-MM-DD"
            )
            all_days[day_idx]["date"] = new_date

            # 逐行编辑
            for row_idx, row in enumerate(rows):
                uncertain_fields = [k.replace("_uncertain", "")
                                    for k in row if k.endswith("_uncertain") and row[k]]

                if uncertain_fields:
                    st.markdown(f"**第 {row.get('seq', '?')} 行** ⚠️ 不确定字段: {', '.join(uncertain_fields)}")
                else:
                    st.markdown(f"**第 {row.get('seq', '?')} 行**")

                cols = st.columns([1, 1, 1.5, 1.5, 1.5, 2, 1.5])
                with cols[0]:
                    row["period"] = st.selectbox(
                        "餐段", ["中", "晚"], key=f"period_{day_idx}_{row_idx}",
                        index=0 if row.get("period") == "中" else 1,
                    )
                with cols[1]:
                    room_val = row.get("room", "")
                    room_options = [""] + VALID_ROOMS
                    try:
                        room_index = room_options.index(room_val)
                    except ValueError:
                        room_options.insert(1, room_val)
                        room_index = 1
                    row["room"] = st.selectbox(
                        "包间号", room_options, index=room_index,
                        key=f"room_{day_idx}_{row_idx}",
                    )
                with cols[2]:
                    row["revenue"] = st.number_input(
                        "营业额", value=float(row.get("revenue") or 0),
                        step=1.0, key=f"rev_{day_idx}_{row_idx}",
                    )
                with cols[3]:
                    row["income"] = st.number_input(
                        "收入", value=float(row.get("income") or 0),
                        step=1.0, key=f"inc_{day_idx}_{row_idx}",
                    )
                with cols[4]:
                    row["actual"] = st.number_input(
                        "实收", value=float(row.get("actual") or 0),
                        step=1.0, key=f"act_{day_idx}_{row_idx}",
                    )
                with cols[5]:
                    pay_val = row.get("payment", "")
                    row["payment"] = st.text_input(
                        "付款方式", value=pay_val,
                        key=f"pay_{day_idx}_{row_idx}",
                        help="多个用/分隔，如: 微信/现金",
                    )
                with cols[6]:
                    pass

            # 备注
            new_notes = st.text_area(
                "备注", value=notes, key=f"notes_{day_idx}",
                height=68,
            )
            all_days[day_idx]["notes"] = new_notes

    if has_any_uncertain:
        st.warning("⚠️ 有标记为不确定的字段（已用黄色标出），请检查后确认。")

    if st.button("✅ 确认无误，生成报表", type="primary"):
        st.session_state.restaurant_results = all_days
        st.session_state.restaurant_confirmed = True
        st.rerun()

# ---------------------------------------------------------------------------
# Step 4 & 5: 生成 Excel & 下载
# ---------------------------------------------------------------------------

if st.session_state.restaurant_confirmed and st.session_state.restaurant_results:
    all_days = st.session_state.restaurant_results

    st.success("正在生成 Excel 报表...")
    excel_bytes = build_restaurant_excel(all_days)

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="📥 下载餐厅日报表",
        data=excel_bytes,
        file_name=f"餐厅日报表_{now_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    if st.button("🔄 重新开始"):
        st.session_state.restaurant_results = None
        st.session_state.restaurant_confirmed = False
        st.rerun()
